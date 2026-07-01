import re
import time
from pathlib import Path
from uuid import UUID

from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Document, DocumentChunk, DocumentScope, LLMUsageLog, Message
from app.core.embeddings import embed_text
from app.services.knowledge_engine import KnowledgeEngine

SYSTEM_PROMPT = """Вы — эксперт по аккредитации испытательных лабораторий, ГОСТ ISO/IEC 17025-2019,
критериям аккредитации, законодательству РФ, требованиям Росаккредитации, метрологии,
неопределённости измерений и внутрилабораторному контролю.

Отвечайте точно, со ссылками на нормативные документы и пункты.
Используйте предоставленный контекст из базы знаний и документов.
Если информации недостаточно — укажите это явно.
Форматируйте ответ в Markdown."""


class DocumentProcessor:
  SUPPORTED = {".pdf", ".docx", ".xlsx", ".txt", ".md", ".html", ".htm"}

  def extract_text(self, file_path: Path, file_type: str) -> list[dict]:
    chunks: list[dict] = []

    if file_type == ".pdf":
      from pypdf import PdfReader

      reader = PdfReader(str(file_path))
      for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
          chunks.append({"content": text.strip(), "index": i, "heading": f"Страница {i + 1}"})

    elif file_type == ".docx":
      from docx import Document as DocxDocument

      doc = DocxDocument(str(file_path))
      current_heading = None
      buffer: list[str] = []
      idx = 0
      for para in doc.paragraphs:
        style = para.style.name if para.style else ""
        if "Heading" in style:
          if buffer:
            chunks.append({"content": "\n".join(buffer), "index": idx, "heading": current_heading})
            idx += 1
            buffer = []
          current_heading = para.text
        elif para.text.strip():
          buffer.append(para.text)
      if buffer:
        chunks.append({"content": "\n".join(buffer), "index": idx, "heading": current_heading})

    elif file_type == ".xlsx":
      from openpyxl import load_workbook

      wb = load_workbook(str(file_path), read_only=True, data_only=True)
      for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
          row_text = " | ".join(str(c) if c is not None else "" for c in row)
          if row_text.strip(" |"):
            rows.append(row_text)
        if rows:
          chunks.append({"content": "\n".join(rows), "index": len(chunks), "heading": sheet_name})

    elif file_type in {".txt", ".md"}:
      text = file_path.read_text(encoding="utf-8", errors="ignore")
      sections = re.split(r"\n(?=#{1,3}\s)", text) if file_type == ".md" else [text]
      for i, section in enumerate(sections):
        heading_match = re.match(r"^#{1,3}\s+(.+)", section)
        heading = heading_match.group(1) if heading_match else None
        chunks.append({"content": section.strip(), "index": i, "heading": heading})

    elif file_type in {".html", ".htm"}:
      from bs4 import BeautifulSoup

      html = file_path.read_text(encoding="utf-8", errors="ignore")
      soup = BeautifulSoup(html, "html.parser")
      for tag in soup.find_all(["h1", "h2", "h3", "p", "li", "td"]):
        text = tag.get_text(strip=True)
        if text:
          heading = tag.name.startswith("h") and text or None
          chunks.append({"content": text, "index": len(chunks), "heading": heading})

    return chunks

  async def index_document(self, db: AsyncSession, document: Document) -> int:
    from app.core.embeddings import embed_text

    file_path = Path(document.file_path)
    file_type = Path(document.filename).suffix.lower()
    raw_chunks = self.extract_text(file_path, file_type)

    count = 0
    for raw in raw_chunks:
      content = raw["content"]
      if len(content) < 20:
        continue
      embedding_resp = embed_text(content[:8000])
      clause_match = re.search(r"(?:п\.|пункт)\s*(\d+(?:\.\d+)*)", content[:200], re.I)
      chunk = DocumentChunk(
        document_id=document.id,
        content=content[:10000],
        chunk_index=raw["index"],
        heading=raw.get("heading"),
        clause=clause_match.group(1) if clause_match else None,
        metadata_={"source_file": document.filename},
        embedding=embedding_resp,
      )
      db.add(chunk)
      count += 1

    document.is_indexed = True
    await db.flush()
    return count


class QuestionProcessor:
  def __init__(self, db: AsyncSession) -> None:
    self.db = db
    self.engine = KnowledgeEngine(db)
    from app.services.llm.provider import get_llm_provider
    self.llm = get_llm_provider()
    self.doc_processor = DocumentProcessor()

  def extract_keywords(self, text: str) -> list[str]:
    words = re.findall(r"[а-яёa-z]{4,}", text.lower())
    stop = {"какие", "какой", "какая", "какое", "котор", "этого", "этом", "быть", "было", "были", "можно", "нужно", "должен"}
    return list(dict.fromkeys(w for w in words if w not in stop))[:15]

  async def process(
    self,
    question: str,
    user_id: UUID,
    chat_id: UUID | None = None,
  ) -> dict:
    start = time.perf_counter()
    topic = self.engine.detect_topic(question)

    cached = await self.engine.find_best_answer(question, owner_id=user_id)
    if cached:
      elapsed = int((time.perf_counter() - start) * 1000)
      return {
        "content": cached.item.content,
        "source": "knowledge_base",
        "relevance_score": cached.score,
        "knowledge_item_id": cached.item.id,
        "used_documents": cached.item.related_documents,
        "tokens_used": 0,
        "response_time_ms": elapsed,
        "topic": topic,
      }

    normalized = self.engine.normalize_text(question)
    embedding = await self.engine.get_embedding(normalized)

    personal_chunks = await self.engine.search_document_chunks(embedding, owner_id=user_id, limit=5)
    global_chunks = await self.engine.search_document_chunks(embedding, owner_id=None, limit=5)
    kb_hits = await self.engine.hybrid_search(question, owner_id=user_id, limit=5)

    context_parts: list[str] = []
    used_docs: list[dict] = []

    for chunk in personal_chunks:
      context_parts.append(f"[Личный документ: {chunk.heading or 'фрагмент'}]\n{chunk.content[:2000]}")
      used_docs.append({"type": "personal", "heading": chunk.heading, "clause": chunk.clause})

    for chunk in global_chunks:
      if chunk not in personal_chunks:
        context_parts.append(f"[Документ: {chunk.heading or 'фрагмент'}]\n{chunk.content[:2000]}")
        used_docs.append({"type": "global", "heading": chunk.heading, "clause": chunk.clause})

    for hit in kb_hits[:3]:
      context_parts.append(
        f"[База знаний: {hit.item.title}]\n{hit.item.content[:1500]}"
      )

    context = "\n\n---\n\n".join(context_parts) if context_parts else "Контекст не найден."
    user_prompt = f"""Контекст:
{context}

Вопрос пользователя:
{question}

Дайте экспертный ответ с указанием нормативных ссылок."""

    llm_start = time.perf_counter()
    llm_response = await self.llm.complete(SYSTEM_PROMPT, user_prompt)
    llm_elapsed = int((time.perf_counter() - llm_start) * 1000)

    keywords = self.extract_keywords(question)
    new_item = await self.engine.create_from_qa(
      question=question,
      answer=llm_response.content,
      category=topic,
      keywords=keywords,
      author_id=user_id,
      source="llm_generated",
      used_documents=used_docs,
    )

    usage_log = LLMUsageLog(
      user_id=user_id,
      provider=self.llm.provider_name,
      model=llm_response.model,
      prompt_tokens=llm_response.prompt_tokens,
      completion_tokens=llm_response.completion_tokens,
      total_tokens=llm_response.total_tokens,
      purpose="question_answer",
    )
    self.db.add(usage_log)

    total_elapsed = int((time.perf_counter() - start) * 1000)
    return {
      "content": llm_response.content,
      "source": "llm_rag",
      "relevance_score": kb_hits[0].score if kb_hits else None,
      "knowledge_item_id": new_item.id,
      "used_documents": used_docs,
      "tokens_used": llm_response.total_tokens,
      "response_time_ms": total_elapsed,
      "topic": topic,
    }
